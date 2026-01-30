import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ConversationHandler, ContextTypes, CallbackQueryHandler
from flask import Flask
from threading import Thread

# --- CONFIGURACIÓN FLASK PARA RENDER ---
server = Flask('')

@server.route('/')
def home():
    return "Bot de Ventas en línea"

def run():
    # Render expone el puerto a través de una variable de entorno PORT
    port = int(os.environ.get('PORT', 8080))
    server.run(host='0.0.0.0', port=port)

def keep_alive():
    t = Thread(target=run)
    t.start()

# --- CONFIGURACIÓN BOT Y EXCEL ---
TOKEN = '8354510389:AAF1_OveG63b9vzW4Ir3nght6dqY0HHg5VE'
EXCEL_PATH = 'datos_clientes.xlsx'
os.makedirs('fotos', exist_ok=True)

# Estructuras de Columnas
COLUMNAS_CLIENTES = ["Rif", "Razón Social", "Numero de Contacto", "Persona contacto", "Numero de Cedula", "Dirección Rif", "Dirección Despacho", "Foto Rif", "Foto Cedula", "Referencia 1", "Referencia 2", "Fecha"]
CAMPOS_REGISTRO = ["Razón Social", "Numero de Contacto", "Persona contacto", "Numero de Cedula", "Dirección Rif", "Dirección Despacho", "Foto Cedula", "Referencia 1", "Referencia 2"]
COLUMNAS_PEDIDOS = ["Código", "Cantidad", "Precio", "Total", "Orden", "Rif", "Razon Social", "Fecha"]

# ESTADOS
ESPERANDO_RIF, REGISTRO_FOTO_RIF, PREGUNTANDO_DATOS, ESPERANDO_UNIDADES, CONTINUAR_O_FINALIZAR = range(5)

# --- FUNCIONES DE EXCEL ---

def obtener_siguiente_orden():
    try:
        if os.path.exists(EXCEL_PATH):
            df_p = pd.read_excel(EXCEL_PATH, sheet_name='Pedidos', engine='openpyxl')
            if not df_p.empty and 'Orden' in df_p.columns:
                return int(df_p['Orden'].max()) + 1
    except: pass
    return 1

def guardar_pedido_inmediato(datos_pedido):
    df_nuevo = pd.DataFrame([datos_pedido]).reindex(columns=COLUMNAS_PEDIDOS)
    if os.path.exists(EXCEL_PATH):
        with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            try:
                book = load_workbook(EXCEL_PATH)
                if 'Pedidos' in book.sheetnames:
                    start_row = book['Pedidos'].max_row
                    df_nuevo.to_excel(writer, sheet_name='Pedidos', index=False, header=False, startrow=start_row)
                else:
                    df_nuevo.to_excel(writer, sheet_name='Pedidos', index=False)
            except:
                df_nuevo.to_excel(writer, sheet_name='Pedidos', index=False)
    else:
        df_nuevo.to_excel(EXCEL_PATH, sheet_name='Pedidos', index=False)

def guardar_cliente_excel(datos):
    datos_finales = {k: datos.get(k, "") for k in COLUMNAS_CLIENTES}
    datos_finales['Fecha'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df_nuevo = pd.DataFrame([datos_finales])
    if os.path.exists(EXCEL_PATH):
        with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            try:
                book = load_workbook(EXCEL_PATH)
                start_row = book['Clientes'].max_row
                df_nuevo.to_excel(writer, sheet_name='Clientes', index=False, header=False, startrow=start_row)
            except:
                df_nuevo.to_excel(writer, sheet_name='Clientes', index=False)

# --- MANEJADORES ---

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 Bienvenido. Ingrese su **RIF** (Ej: J123456789):")
    return ESPERANDO_RIF

async def validar_rif(update: Update, context: ContextTypes.DEFAULT_TYPE):
    rif_usuario = update.message.text.upper().strip()
    context.user_data['Rif'] = rif_usuario
    context.user_data['Orden'] = obtener_siguiente_orden()

    try:
        if os.path.exists(EXCEL_PATH):
            df_cli = pd.read_excel(EXCEL_PATH, sheet_name='Clientes', dtype=str, engine='openpyxl')
            df_cli.columns = df_cli.columns.str.strip()
            cliente = df_cli[df_cli['Rif'].str.upper() == rif_usuario]
            if not cliente.empty:
                context.user_data['Razon Social'] = cliente.iloc[0]['Razón Social']
                await update.message.reply_text(f"✅ Cliente: {context.user_data['Razon Social']}")
                await mostrar_menu_productos(update, context, pagina=0)
                return ESPERANDO_UNIDADES
    except Exception as e:
        print(f"Error: {e}")
    
    await update.message.reply_text("🔍 No registrado. Envíe **Foto del RIF**:")
    return REGISTRO_FOTO_RIF

# --- FUNCIONES AÑADIDAS PARA CONVERSATION HANDLER ---
async def guardar_foto_rif(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Lógica placeholder: guarda la foto y pide el siguiente dato
    photo_file = await update.message.photo[-1].get_file()
    path = os.path.join('fotos', f"{context.user_data['Rif']}_rif.jpg")
    await photo_file.download_to_drive(path)
    context.user_data['Foto Rif'] = path
    await update.message.reply_text("Foto del RIF guardada. Ingrese Razón Social:")
    context.user_data['campo_actual'] = 0 # Usamos un índice para CAMPOS_REGISTRO
    return PREGUNTANDO_DATOS

async def registro_cliente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Lógica placeholder: guarda datos de registro
    campo = CAMPOS_REGISTRO[context.user_data['campo_actual']]
    context.user_data[campo] = update.message.text
    
    if context.user_data['campo_actual'] < len(CAMPOS_REGISTRO) - 1:
        context.user_data['campo_actual'] += 1
        siguiente_campo = CAMPOS_REGISTRO[context.user_data['campo_actual']]
        await update.message.reply_text(f"Ingrese {siguiente_campo}:")
        return PREGUNTANDO_DATOS
    else:
        guardar_cliente_excel(context.user_data)
        await update.message.reply_text("✅ Registro completado. Proceda con el pedido.")
        await mostrar_menu_productos(update, context, pagina=0)
        return ESPERANDO_UNIDADES
# --- FIN FUNCIONES AÑADIDAS ---


async def mostrar_menu_productos(update: Update, context: ContextTypes.DEFAULT_TYPE, pagina: int = 0):
    df_prod = pd.read_excel(EXCEL_PATH, sheet_name='Productos', engine='openpyxl')
    items_por_pagina = 10
    inicio, fin = pagina * items_por_pagina, (pagina + 1) * items_por_pagina
    df_pag = df_prod.iloc[inicio:fin]
    
    keyboard = []
    for _, row in df_pag.iterrows():
        texto_boton = f"{row['Descripcion']} - ${row['Precio']}"
        keyboard.append([InlineKeyboardButton(texto_boton, callback_data=f"p_{row['Código']}")])

    nav = []
    if pagina > 0: nav.append(InlineKeyboardButton("⬅️ Ant.", callback_data=f"pag_{pagina-1}"))
    if fin < len(df_prod): nav.append(InlineKeyboardButton("Sig. ➡️", callback_data=f"pag_{pagina+1}"))
    if nav: keyboard.append(nav)

    markup = InlineKeyboardMarkup(keyboard)
    msg = f"📦 **Catálogo (Pág. {pagina+1})**\nSeleccione un producto:"
    if update.callback_query:
        await update.callback_query.edit_message_text(msg, reply_markup=markup, parse_mode='Markdown')
    else:
        await update.message.reply_text(msg, reply_markup=markup, parse_mode='Markdown')

async def seleccion_producto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data.startswith("pag_"):
        await mostrar_menu_productos(update, context, pagina=int(query.data.split("_")[1]))
        return ESPERANDO_UNIDADES

    cod = query.data.split("_")[1]
    df_p = pd.read_excel(EXCEL_PATH, sheet_name='Productos', engine='openpyxl')
    p = df_p[df_p['Código'].astype(str) == cod].iloc[0]
    context.user_data['item'] = {'Código': p['Código'], 'Descripcion': p['Descripcion'], 'Precio': p['Precio']}
    await query.edit_message_text(f"📦 **{p['Descripcion']}**\nPrecio: ${p['Precio']}\n\n¿Cuántas unidades desea?")
    return ESPERANDO_UNIDADES

async def recibir_unidades(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message.text.isdigit(): return ESPERANDO_UNIDADES
    cant = int(update.message.text)
    item = context.user_data['item']
    num_orden = context.user_data['Orden']
    
    pedido = {
        'Código': item['Código'], 'Cantidad': cant, 'Precio': item['Precio'],
        'Total': cant * item['Precio'], 'Orden': num_orden,
        'Rif': context.user_data['Rif'], 'Razon Social': context.user_data.get('Razon Social', 'Nuevo'),
        'Fecha': datetime.now().strftime("%d/%m/%Y %H:%M")
    }
    guardar_pedido_inmediato(pedido)
    
    # Cálculo de subtotal acumulado
    subtotal_orden = pedido['Total']
    try:
        df_hist = pd.read_excel(EXCEL_PATH, sheet_name='Pedidos', engine='openpyxl')
        subtotal_orden = df_hist[df_hist['Orden'] == num_orden]['Total'].sum()
    except: pass

    kyb = InlineKeyboardMarkup([[InlineKeyboardButton("➕ Otro", callback_data="pag_0"), 
                                 InlineKeyboardButton("✅ Fin", callback_data="fin")]])
    await update.message.reply_text(
        f"✅ **Añadido:** {item['Descripcion']}\n"
        f"💰 Total ítem: ${pedido['Total']:.2f}\n"
        f"📋 **Subtotal Orden #{num_orden}: ${subtotal_orden:.2f}**",
        reply_markup=kyb, parse_mode='Markdown'
    )
    return CONTINUAR_O_FINALIZAR

async def finalizar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data.startswith("pag_"):
        await mostrar_menu_productos(update, context, 0)
        return ESPERANDO_UNIDADES
    await query.edit_message_text(f"🏁 Orden #{context.user_data['Orden']} finalizada. ¡Gracias!")
    context.user_data.clear() # Limpia los datos de usuario al finalizar
    return ConversationHandler.END


# --- MAIN FUNCTION CON FLASK KEEP_ALIVE ---

def main():
    keep_alive() # Esto inicia el servidor web para que Render no lo apague
    app = Application.builder().token(TOKEN).connect_timeout(90).read_timeout(90).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ESPERANDO_RIF:,
            REGISTRO_FOTO_RIF: [MessageHandler(filters.PHOTO & ~filters.COMMAND, guardar_foto_rif)],
            PREGUNTANDO_DATOS:,
            ESPERANDO_UNIDADES:,
            CONTINUAR_O_FINALIZAR: [CallbackQueryHandler(finalizar)],
        },
        fallbacks=[CommandHandler('start', start)]
    )

    app.add_handler(conv_handler)
    
    # Inicia el bot en modo polling 
    app.run_polling()

if __name__ == '__main__':
    main()
